#!/usr/bin/env python3
"""
Générateur de fichiers XML de configuration pour Cisco 7942 (SIP)
Usage : python3 generate_cisco_xml.py ListeTelephones.xlsx
 
Règles :
  - Numéro de poste = numéro extrait du nom de salle (X113 -> 113)
  - Nom du fichier  = SEP{MAC_EN_MAJUSCULES}.cnf.xml
  - MAC en majuscules dans le nom de fichier uniquement
  - Mot de passe extension fixe : 123456789
  - Le script génère les fichiers XML dans un dossier "output/" à côté du script
"""
 
# ─── IMPORTS ──────────────────────────────────────────────────────────────────
# sys  : récupération des arguments passés en ligne de commande (nom du fichier Excel)
# os   : gestion des chemins de fichiers et création du dossier de sortie
# re   : expressions régulières pour extraire le numéro depuis le nom de salle (ex: X113 → 113)
# openpyxl : lecture du fichier Excel sans avoir besoin d'installer Microsoft Excel
import sys
import os
import re
import openpyxl
 
# ─── CONFIGURATION GLOBALE ────────────────────────────────────────────────────
# Tous les paramètres réseau et d'authentification sont centralisés ici.
# Modifier uniquement ce bloc pour adapter le script à un autre environnement.
FREEPBX_IP    = "192.168.10.1"   # Adresse IP du serveur FreePBX / TFTP sur le VLAN 10
SSH_USER      = "cisco"          # Identifiant SSH embarqué dans le XML (accès admin du téléphone)
SSH_PASSWORD  = "cisco29200"     # Mot de passe SSH du téléphone
PHONE_PASSWORD = "29200"         # Mot de passe du menu local du téléphone
AUTH_PASSWORD  = "123456789"     # Mot de passe d'authentification SIP (identique pour tous les postes)
 
# ─── TEMPLATE XML ─────────────────────────────────────────────────────────────
# Modèle de fichier de configuration attendu par le firmware SIP42.9-4-2ES26
# du Cisco 7942. Les champs entre accolades ({extension}, {freepbx_ip}, etc.)
# sont des variables Python qui seront remplacées dynamiquement par la fonction
# generer_xml() via la méthode str.format().
# Ce template est chargé une seule fois en mémoire et réutilisé pour chaque téléphone.
TEMPLATE = """<device>
    <deviceProtocol>SIP</deviceProtocol>
    <sshUserId>{ssh_user}</sshUserId>
    <sshPassword>{ssh_password}</sshPassword>
    <ipAddressMode>0</ipAddressMode>
 
    <devicePool>
        <dateTimeSetting>
            <dateTemplate>D/M/Ya</dateTemplate>
            <timeZone>GMT Standard/Daylight Time</timeZone>
            <olsonTimeZone>Europe/Paris</olsonTimeZone>
            <ntps>
                <ntp>
                    <name>145.238.80.83</name>
                    <ntpMode>Unicast</ntpMode>
                </ntp>
            </ntps>
        </dateTimeSetting>
 
        <callManagerGroup>
            <members>
                <member priority="0">
                    <callManager>
                        <ports>
                            <ethernetPhonePort>2000</ethernetPhonePort>
                            <sipPort>5060</sipPort>
                        </ports>
                        <processNodeName>{freepbx_ip}</processNodeName>
                    </callManager>
                </member>
            </members>
        </callManagerGroup>
    </devicePool>
 
    <sipProfile>
        <sipProxies>
            <registerWithProxy>true</registerWithProxy>
        </sipProxies>
        <sipCallFeatures>
            <cnfJoinEnabled>true</cnfJoinEnabled>
            <rfc2543Hold>false</rfc2543Hold>
            <callHoldRingback>2</callHoldRingback>
            <localCfwdEnable>true</localCfwdEnable>
            <semiAttendedTransfer>true</semiAttendedTransfer>
            <anonymousCallBlock>2</anonymousCallBlock>
            <callerIdBlocking>2</callerIdBlocking>
            <dndControl>0</dndControl>
            <remoteCcEnable>true</remoteCcEnable>
            <callForwardURI>x-serviceuri-cfwdall</callForwardURI>
            <callPickupURI>x-cisco-serviceuri-pickup</callPickupURI>
            <callPickupListURI>x-cisco-serviceuri-opickup</callPickupListURI>
            <callPickupGroupURI>x-cisco-serviceuri-gpickup</callPickupGroupURI>
            <meetMeServiceURI>x-cisco-serviceuri-meetme</meetMeServiceURI>
            <abbreviatedDialURI>x-cisco-serviceuri-abbrdial</abbreviatedDialURI>
        </sipCallFeatures>
 
        <sipStack>
            <sipInviteRetx>6</sipInviteRetx>
            <sipRetx>10</sipRetx>
            <timerInviteExpires>180</timerInviteExpires>
            <timerRegisterExpires>3600</timerRegisterExpires>
            <timerRegisterDelta>5</timerRegisterDelta>
            <timerKeepAliveExpires>120</timerKeepAliveExpires>
            <timerSubscribeExpires>120</timerSubscribeExpires>
            <timerSubscribeDelta>5</timerSubscribeDelta>
            <timerT1>500</timerT1>
            <timerT2>4000</timerT2>
            <maxRedirects>70</maxRedirects>
            <remotePartyID>true</remotePartyID>
            <userInfo>None</userInfo>
        </sipStack>
 
        <autoAnswerTimer>1</autoAnswerTimer>
        <autoAnswerAltBehavior>false</autoAnswerAltBehavior>
        <autoAnswerOverride>true</autoAnswerOverride>
        <transferOnhookEnabled>false</transferOnhookEnabled>
        <enableVad>false</enableVad>
        <preferredCodec>g711alaw</preferredCodec>
        <dtmfAvtPayload>101</dtmfAvtPayload>
        <dtmfDbLevel>3</dtmfDbLevel>
        <dtmfOutofBand>avt</dtmfOutofBand>
        <alwaysUsePrimeLine>false</alwaysUsePrimeLine>
        <alwaysUsePrimeLineVoiceMail>false</alwaysUsePrimeLineVoiceMail>
        <kpml>3</kpml>
        <natEnabled>false</natEnabled>
        <phoneLabel>{extension}</phoneLabel>
        <stutterMsgWaiting>0</stutterMsgWaiting>
        <callStats>false</callStats>
        <silentPeriodBetweenCallWaitingBursts>10</silentPeriodBetweenCallWaitingBursts>
        <disableLocalSpeedDialConfig>false</disableLocalSpeedDialConfig>
        <startMediaPort>10000</startMediaPort>
        <stopMediaPort>20000</stopMediaPort>
 
        <sipLines>
            <line button="1">
                <featureID>9</featureID>
                <featureLabel>{feature_label}</featureLabel>
                <proxy>USECALLMANAGER</proxy>
                <port>5060</port>
                <name>{extension}</name>
                <displayName>{extension}</displayName>
                <autoAnswer>
                    <autoAnswerEnabled>2</autoAnswerEnabled>
                </autoAnswer>
                <callWaiting>3</callWaiting>
                <authName>{extension}</authName>
                <authPassword>{auth_password}</authPassword>
                <sharedLine>false</sharedLine>
                <messageWaitingLampPolicy>1</messageWaitingLampPolicy>
                <messagesNumber>*97</messagesNumber>
                <ringSettingIdle>4</ringSettingIdle>
                <ringSettingActive>5</ringSettingActive>
                <contact>{extension}</contact>
                <forwardCallInfoDisplay>
                    <callerName>true</callerName>
                    <callerNumber>true</callerNumber>
                    <redirectedNumber>false</redirectedNumber>
                    <dialedNumber>true</dialedNumber>
                </forwardCallInfoDisplay>
            </line>
        </sipLines>
 
        <voipControlPort>5060</voipControlPort>
        <dscpForAudio>184</dscpForAudio>
        <ringSettingBusyStationPolicy>0</ringSettingBusyStationPolicy>
        <dialTemplate>dialplan.xml</dialTemplate>
    </sipProfile>
 
    <commonProfile>
        <phonePassword>{phone_password}</phonePassword>
        <backgroundImageAccess>true</backgroundImageAccess>
        <callLogBlfEnabled>1</callLogBlfEnabled>
    </commonProfile>
 
    <loadInformation>SIP42.9-4-2ES26</loadInformation>
 
    <vendorConfig>
        <disableSpeaker>false</disableSpeaker>
        <disableSpeakerAndHeadset>false</disableSpeakerAndHeadset>
        <pcPort>0</pcPort>
        <settingsAccess>1</settingsAccess>
        <garp>0</garp>
        <voiceVlanAccess>0</voiceVlanAccess>
        <videoCapability>0</videoCapability>
        <autoSelectLineEnable>0</autoSelectLineEnable>
        <webAccess>0</webAccess>
        <spanToPCPort>1</spanToPCPort>
        <loggingDisplay>1</loggingDisplay>
        <loadServer>{freepbx_ip}</loadServer>
        <sshAccess>0</sshAccess>
        <sshPort>22</sshPort>
    </vendorConfig>
 
    <versionStamp>002</versionStamp>
    <networkLocale>France</networkLocale>
    <networkLocaleInfo>
        <name>France</name>
        <uid>11</uid>
        <version>1.0.0.0-4</version>
    </networkLocaleInfo>
 
    <deviceSecurityMode>0</deviceSecurityMode>
    <authenticationURL></authenticationURL>
    <servicesURL></servicesURL>
    <directoryURL>http://{freepbx_ip}/directory.xml</directoryURL>
    <idleURL></idleURL>
    <informationURL></informationURL>
    <messagesURL></messagesURL>
    <proxyServerURL></proxyServerURL>
    <dialToneSetting>2</dialToneSetting>
    <dscpForSCCPPhoneConfig>96</dscpForSCCPPhoneConfig>
    <dscpForSCCPPhoneServices>0</dscpForSCCPPhoneServices>
    <dscpForCm2Dvce>96</dscpForCm2Dvce>
    <capfAuthMode>0</capfAuthMode>
    <capfList>
        <capf>
            <phonePort>3804</phonePort>
        </capf>
    </capfList>
    <transportLayerProtocol>2</transportLayerProtocol>
    <certHash></certHash>
    <encrConfig>false</encrConfig>
</device>
"""
 
 
def extraire_numero(nom_salle):
    """
    Extrait le numéro d'extension depuis le nom de salle.
    Utilise une expression régulière pour isoler la première suite de chiffres trouvée.
    Exemples : "X113" → "113" / "X121" → "121" / "salle 105" → "105"
    Retourne None si aucun chiffre n'est trouvé (ligne ignorée par generer_xml).
    """
    match = re.search(r'\d+', str(nom_salle))
    if match:
        return match.group()
    return None
 
 
def generer_xml(mac, nom_salle, output_dir):
    """
    Génère le fichier XML de configuration pour un téléphone Cisco 7942.
 
    Paramètres :
      mac        : adresse MAC du téléphone (ex: "08CC68E9690D")
      nom_salle  : nom de la salle lu depuis l'Excel (ex: "X113")
      output_dir : chemin du dossier de sortie où écrire le fichier
 
    Le nom du fichier généré suit la convention imposée par le firmware Cisco :
    SEP{ADRESSE_MAC_EN_MAJUSCULES}.cnf.xml
    Ce fichier est déposé dans le dossier TFTP de FreePBX pour être téléchargé
    automatiquement par le téléphone au démarrage.
    """
    # Extraction du numéro d'extension depuis le nom de salle
    extension = extraire_numero(nom_salle)
    if not extension:
        print(f"  [SKIP] Impossible d'extraire le numéro depuis '{nom_salle}'")
        return
 
    # Mise en forme de l'adresse MAC et du label affiché sur l'écran du téléphone
    mac_upper     = mac.strip().upper()
    feature_label = nom_salle.strip().upper()
 
    # Construction du nom de fichier imposé par Cisco : SEP + MAC en majuscules + .cnf.xml
    filename  = f"SEP{mac_upper}.cnf.xml"
    filepath  = os.path.join(output_dir, filename)
 
    # Injection des variables dans le template XML et écriture du fichier en UTF-8
    content = TEMPLATE.format(
        ssh_user=SSH_USER,
        ssh_password=SSH_PASSWORD,
        freepbx_ip=FREEPBX_IP,
        extension=extension,
        feature_label=feature_label,
        auth_password=AUTH_PASSWORD,
        phone_password=PHONE_PASSWORD,
    )
 
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
 
    print(f"  [OK] {filename}  ->  salle {nom_salle}, poste {extension}")
 
 
def main():
    """
    Point d'entrée du script.
    Vérifie les arguments, charge le fichier Excel, détecte automatiquement
    les colonnes 'salle' et 'mac', puis appelle generer_xml() pour chaque ligne.
    """
    # Vérification qu'un fichier Excel a bien été passé en argument
    if len(sys.argv) < 2:
        print("Usage : python3 generate_cisco_xml.py <fichier.xlsx>")
        sys.exit(1)
 
    xlsx_file = sys.argv[1]
 
    if not os.path.isfile(xlsx_file):
        print(f"Erreur : fichier introuvable : {xlsx_file}")
        sys.exit(1)
 
    # Création du dossier de sortie "output/" à côté du fichier Excel
    script_dir = os.path.dirname(os.path.abspath(xlsx_file))
    output_dir = os.path.join(script_dir, "output")
    os.makedirs(output_dir, exist_ok=True)
 
    print(f"\nLecture de : {xlsx_file}")
    print(f"Dossier de sortie : {output_dir}\n")
 
    # Chargement du fichier Excel en mode lecture seule pour limiter la mémoire utilisée
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    ws = wb.active
 
    # Détection automatique des colonnes depuis la ligne d'en-tête (ligne 1)
    # Permet de fonctionner quel que soit l'ordre des colonnes dans le fichier Excel
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
 
    try:
        col_salle = headers.index(next(h for h in headers if "salle" in h or "num" in h))
        col_mac   = headers.index(next(h for h in headers if "mac" in h))
    except StopIteration:
        print("Erreur : colonnes 'salle' ou 'mac' introuvables dans le fichier Excel.")
        print(f"Colonnes détectées : {headers}")
        sys.exit(1)
 
    # Parcours des lignes de données (à partir de la ligne 2) et génération des XML
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nom_salle = row[col_salle]
        mac       = row[col_mac]
 
        # Ignore les lignes vides ou incomplètes
        if not nom_salle or not mac:
            continue
 
        generer_xml(str(mac), str(nom_salle), output_dir)
        count += 1
 
    print(f"\n{count} fichier(s) XML généré(s) dans : {output_dir}")
 
 
# Point d'entrée standard Python : n'exécute main() que si le script
# est lancé directement (et non importé comme module dans un autre script)
if __name__ == "__main__":
    main()
